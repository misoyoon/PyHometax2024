import openpyxl


def read_from_file(filepath):
    wb = openpyxl.load_workbook(filename=filepath)
    ws = wb.active

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        value_list = []
        for cell in row:
            if cell.is_date:
                value = cell.value.strftime('%Y-%m-%d')
            else:
                value = cell.value if cell.value else ''
            value_list.append('{:>15}'.format(value))

        print(''.join(value_list))

    wb.close()


if __name__ == '__main__':
    read_from_file('excel_data\\test.xlsx')