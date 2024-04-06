import sys, os
import math
import decimal
import datetime
from typing import List, Dict, Set, Final, Union, Any


from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

import sys
import pymysql

import config
import common
import ht_file

def set_global(group_id, v_host_name, v_user_id, v_ht_tt_seq, v_au_step):
    global host_name
    host_name = v_host_name
    global user_id
    user_id = v_user_id
    global ht_tt_seq
    ht_tt_seq = v_ht_tt_seq
    global au_step
    au_step = v_au_step

def connect_db():
    #if env != None:
    #    raise ValueError("환경정보를 설정해주세요")
    global conn    
    conn = pymysql.connect(
                            db  =config.DATABASE_CONFIG['db'],
                            user=config.DATABASE_CONFIG['user'],
                            host=config.DATABASE_CONFIG['host'],
                            port=config.DATABASE_CONFIG['port'],
                            password=config.DATABASE_CONFIG['password']
                        )
    return conn

conn = None
set_global(group_id, )
connect_db()


def insert_auditWt(import_list: List[Dict[str, str]]) :
    # Connection 으로부터 Cursor 생성
    curs = conn.cursor(pymysql.cursors.DictCursor)

    # set session size
    #sql ='set global max_allowed_packet=25*1024*1024;'
    #curs.execute(sql)

    # SQL문 실행
    sql = '''
        INSERT INTO ht_tt(
            ht_series_yyyymm,
            step_cd,
            series_seq,
            holder_nm,
            main_account,
            total_income_amount,
            holder_full_addr,
            holder_cellphone,
            notify_type_cd,
            holder_email,
            data_type,
            sec_company_cd,
            sec_branch_nm,
            sec_charger_nm,
            holder_ssn1,
            holder_ssn2,
            reg_id,
            import_seq,
            remark
        ) VALUES (
            %(ht_series_yyyymm)s,
            %(step_cd)s,
            %(series_seq)s, 
            %(holder_nm)s,
            %(main_account)s,
            %(total_income_amount)s,
            %(holder_full_addr)s,
            %(holder_cellphone)s,
            %(notify_type_cd)s,
            %(holder_email)s,
            %(data_type)s,
            %(sec_company_cd)s,
            %(sec_branch_nm)s,
            %(sec_charger_nm)s,
            %(holder_ssn1)s,
            %(holder_ssn2)s,
            %(reg_id)s,
            %(import_seq)s,
            %(remark)s
        )

    '''
    affectedTotal = 0
    for param in import_list:
        try:
            #logqry(sql, param)
            affected = curs.execute(sql, param)
            affectedTotal += affected
            if affected == 0:
                print(f'Insert 실패 = {param}')
                
        except Exception as e:
            if f'{e}'.find('Duplicate') >= 0:
                ...
            else:
                print(f'insert_HtTt() ERROR : {e}')
            #print(param)
            
    conn.commit()   
    return affectedTotal


def to_col(ch: str):
    ch = ch.upper()
    return ord(ch) - 64

def to_str(num: Any) -> str:
    if num:
        return str(num).strip()
    else:
        return ''

def to_num(num: str) -> int:
    if num:
        return int(str(num).strip().replace(',', ''))
    else:
        return 0

mapping = {
    '결과':'A',
    '세목':'B',
    '양도인명':'C',
    '세액':'D',
    '납기일':'E',
    '신고일':'F',
    '주소':'G',
    '위택스접수번호':'H',
}

def read_excel(excel_file:str, next_seq:int, next_import_seq:int):
    
    #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = load_workbook(excel_file, data_only=True)
    #시트 이름으로 불러오기
    ws = load_wb['Sheet1']

    insert_seq = next_seq;

    import_list: List[ Dict[str, str]] = []
    #for row in range(2, 5):
    for row in range(2, ws.max_row+1):
        rowData = {}

        try:
            결과        = to_str(ws.cell(row, 1).value)
            세목        = to_num(ws.cell(row, 2).value)
            양도인명    = to_str(ws.cell(row, 3).value)
            세액        = to_str(ws.cell(row, 4).value)
            납기일      = to_str(ws.cell(row, 5).value)
            신고일      = to_str(ws.cell(row, 6).value)
            주소        = to_str(ws.cell(row, 7).value)          # NULL 가능 (strip사용불가능)
            접수번호    = to_str(ws.cell(row, 8).value)         # NULL 가능 (strip사용불가능)
        
            rowData['pay_type'] = 결과
            rowData['tax_type'] = 세목
            rowData['holder_nm'] = 양도인명
            rowData['tax'] = 세액
            rowData['pay_to_date'] = 납기일       
            rowData['wetax_reg_dt'] = 신고일
            rowData['si_do'] = 주소   
            rowData['holder_ssn1'] = 접수번호   
            insert_seq += 1

            #print(rowData)
            import_list.append(rowData)

        except Exception as e:
            print(f'엑셀 데이터 로딩시 오류 발생: {e} - Row:{row}, 양도인명={holder_nm}')

    # 데이터 반입 실행
    #print(import_list[671])
    affectedTotal = insert_HtTt(import_list)
    #affectedTotal = update_HtTt(import_list)
    print('----------------------------------------------')
    print(f'반입결과 : 총={len(import_list)}, 반입 성공={affectedTotal} , 실패건수={len(import_list) - affectedTotal}')



def main():
    excel_file = 'C:/Users/rebor/OneDrive/문서/위택스_신고내역_20240515_1050.xlsx'

    rs_seq = get_HtTt_NextSeq(HT_SERIES_YYYYMM)
    import_seq = get_next_반입회차(HT_SERIES_YYYYMM)
    
    if os.path.exists(excel_file):
        read_excel(excel_file, rs_seq, import_seq)
    else:
        print(f'리스트 엑셀 파일을 찾을 수 없습니다. {excel_file}')


# 메인함수 실행
main()